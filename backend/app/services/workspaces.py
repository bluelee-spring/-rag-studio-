from __future__ import annotations

import csv
import html
import io
import json
import re
import shutil
import sqlite3
import threading
import unicodedata
import uuid
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from html.parser import HTMLParser
from itertools import islice
from pathlib import Path
from typing import Any, Callable

from ..config import settings
from ..data import TeachingData
from ..models import IngestionJobResponse, WorkspaceInfo


DOCUMENT_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".html", ".htm"}
TABLE_EXTENSIONS = {".csv", ".xlsx"}
GRAPH_EXTENSIONS = {".json", ".jsonl"}
MAX_FILE_BYTES = 25 * 1024 * 1024
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_DOCUMENT_CHARACTERS_PER_FILE = 2_000_000
MAX_DOCUMENT_CHARACTERS_TOTAL = 5_000_000
MAX_DOCUMENT_CHUNKS = 10_000
MAX_TABLE_ROWS = 200_000
MAX_TABLE_COLUMNS = 300


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False) + "\n" for row in rows
        ),
        encoding="utf-8",
    )
    temporary.replace(path)


def _safe_upload_name(value: str) -> str:
    name = Path(value.replace("\\", "/")).name.strip()
    name = re.sub(r"[\x00-\x1f<>:\"/\\|?*]", "_", name)
    return name[:180] or "upload"


def _workspace_id(kind: str) -> str:
    return f"ws-{kind[:3]}-{uuid.uuid4().hex[:10]}"


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        value = html.unescape(data).strip()
        if value:
            self.parts.append(value)


def _extract_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"无法识别文本编码：{path.name}")
    if suffix in {".html", ".htm"}:
        source = path.read_text(encoding="utf-8", errors="replace")
        parser = _TextExtractor()
        parser.feed(source)
        return "\n".join(parser.parts)
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("读取PDF需要安装pypdf") from exc
        reader = PdfReader(str(path))
        return "\n\n".join((page.extract_text() or "") for page in reader.pages)
    if suffix == ".docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise RuntimeError("读取DOCX需要安装python-docx") from exc
        document = Document(str(path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    raise ValueError(f"不支持的文档格式：{suffix}")


def _normalize_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def _split_long_piece(piece: str, size: int) -> list[str]:
    if len(piece) <= size:
        return [piece]
    sentences = re.split(r"(?<=[。！？!?；;\.])\s*", piece)
    result: list[str] = []
    buffer = ""
    for sentence in sentences:
        if not sentence:
            continue
        if len(sentence) > size:
            if buffer:
                result.append(buffer)
                buffer = ""
            result.extend(
                sentence[index : index + size]
                for index in range(0, len(sentence), size)
            )
        elif len(buffer) + len(sentence) <= size:
            buffer += sentence
        else:
            result.append(buffer)
            buffer = sentence
    if buffer:
        result.append(buffer)
    return result


def chunk_text(text: str, target_size: int, overlap: int) -> list[str]:
    """优先沿段落和句子切分；只有超长句才按字符窗口切分。"""
    normalized = _normalize_text(text)
    if not normalized:
        return []
    if len(normalized) <= target_size:
        return [normalized]
    pieces: list[str] = []
    for paragraph in re.split(r"\n\s*\n", normalized):
        pieces.extend(_split_long_piece(paragraph.strip(), target_size))

    chunks: list[str] = []
    buffer = ""
    for piece in pieces:
        separator = "\n\n" if buffer else ""
        if len(buffer) + len(separator) + len(piece) <= target_size:
            buffer += separator + piece
            continue
        if buffer:
            chunks.append(buffer.strip())
        prefix = buffer[-overlap:] if overlap and buffer else ""
        buffer = (prefix + ("\n" if prefix else "") + piece).strip()
        if len(buffer) > target_size + overlap:
            chunks.extend(_split_long_piece(buffer, target_size))
            buffer = ""
    if buffer:
        chunks.append(buffer.strip())
    return [item for item in chunks if item]


def _tokenize(text: str, stopwords: set[str]) -> list[str]:
    normalized = unicodedata.normalize("NFKC", text).lower()
    try:
        import jieba

        candidates = jieba.lcut(normalized, cut_all=False)
    except ImportError:
        candidates = re.findall(
            r"[a-z][a-z0-9_+.-]*|\d+(?:\.\d+)?|[\u4e00-\u9fff]{1,4}",
            normalized,
        )
    tokens = []
    for value in candidates:
        token = value.strip()
        if not token or token in stopwords:
            continue
        if re.fullmatch(r"[^\w\u4e00-\u9fff]+", token):
            continue
        tokens.append(token)
    return tokens


def _read_csv(path: Path) -> tuple[list[str], list[list[Any]], str]:
    decoded = ""
    used_encoding = ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            decoded = path.read_text(encoding=encoding)
            used_encoding = encoding
            break
        except UnicodeDecodeError:
            continue
    if not decoded:
        raise ValueError("CSV为空或编码无法识别")
    sample = decoded[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(decoded, newline=""), dialect=dialect))
    if not rows:
        raise ValueError("CSV没有可读取的数据行")
    width = max(len(row) for row in rows)
    if width > MAX_TABLE_COLUMNS:
        raise ValueError(f"表格列数不能超过{MAX_TABLE_COLUMNS}")
    headers = [
        (rows[0][index].strip() if index < len(rows[0]) else "")
        or f"column_{index + 1}"
        for index in range(width)
    ]
    values = [
        [row[index] if index < len(row) else None for index in range(width)]
        for row in rows[1:]
        if any(str(value).strip() for value in row)
    ]
    if len(values) > MAX_TABLE_ROWS:
        raise ValueError(f"表格数据行不能超过{MAX_TABLE_ROWS}")
    return headers, values, used_encoding


def _read_excel(
    path: Path, sheet_name: str
) -> tuple[list[str], list[list[Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取Excel需要安装openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"工作表不存在；可选值：{', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        rows = list(
            islice(
                sheet.iter_rows(values_only=True),
                MAX_TABLE_ROWS + 2,
            )
        )
        if not rows:
            raise ValueError("Excel工作表为空")
        if len(rows) > MAX_TABLE_ROWS + 1:
            raise ValueError(f"表格数据行不能超过{MAX_TABLE_ROWS}")
        width = max(len(row) for row in rows)
        if width > MAX_TABLE_COLUMNS:
            raise ValueError(f"表格列数不能超过{MAX_TABLE_COLUMNS}")
        headers = [
            str(rows[0][index]).strip()
            if index < len(rows[0]) and rows[0][index] is not None
            else f"column_{index + 1}"
            for index in range(width)
        ]
        values = [
            [
                row[index] if index < len(row) else None
                for index in range(width)
            ]
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]
        return headers, values, sheet.title
    finally:
        workbook.close()


def _infer_type(values: list[Any]) -> str:
    nonempty = [value for value in values if value not in (None, "")]
    if not nonempty:
        return "TEXT"
    integer = True
    real = True
    for value in nonempty:
        if isinstance(value, bool):
            integer = real = False
            break
        try:
            numeric = float(value)
            if not numeric.is_integer():
                integer = False
        except (TypeError, ValueError):
            integer = real = False
            break
    if integer:
        return "INTEGER"
    if real:
        return "REAL"
    return "TEXT"


def _coerce(value: Any, sql_type: str) -> Any:
    if value in (None, ""):
        return None
    if sql_type == "INTEGER":
        return int(float(value))
    if sql_type == "REAL":
        return float(value)
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return str(value)


def _unique_business_headers(headers: list[str]) -> list[tuple[str, str]]:
    """返回(展示名, 原始名)，避免重复表头在JSON结果中相互覆盖。"""
    counts: Counter[str] = Counter()
    result: list[tuple[str, str]] = []
    for original in headers:
        key = unicodedata.normalize("NFKC", original).casefold()
        counts[key] += 1
        display = (
            original
            if counts[key] == 1
            else f"{original}（{counts[key]}）"
        )
        result.append((display, original))
    return result


class WorkspaceRegistry:
    def __init__(self, root: Path | None = None) -> None:
        self.root = root or settings.runtime_dir / "workspaces"
        self.root.mkdir(parents=True, exist_ok=True)

    def _path(self, workspace_id: str) -> Path:
        if not re.fullmatch(r"ws-[a-z]{3}-[a-f0-9]{10}", workspace_id):
            raise KeyError("工作区ID无效")
        path = self.root / workspace_id
        if not path.is_dir():
            raise KeyError("工作区不存在")
        return path

    def list(self) -> list[WorkspaceInfo]:
        result: list[WorkspaceInfo] = []
        for manifest_path in self.root.glob("ws-*/workspace.json"):
            try:
                result.append(
                    WorkspaceInfo.model_validate_json(
                        manifest_path.read_text(encoding="utf-8")
                    )
                )
            except (ValueError, OSError):
                continue
        return sorted(result, key=lambda item: item.updated_at, reverse=True)

    def get(self, workspace_id: str) -> WorkspaceInfo:
        path = self._path(workspace_id) / "workspace.json"
        return WorkspaceInfo.model_validate_json(
            path.read_text(encoding="utf-8")
        )

    def load_document_data(self, workspace_id: str) -> TeachingData:
        workspace = self._path(workspace_id)
        manifest = self.get(workspace_id)
        if manifest.kind != "documents" or manifest.status != "ready":
            raise ValueError("该工作区不是可查询的文档工作区")
        documents = workspace / "documents"
        chunks = [
            json.loads(line)
            for line in (documents / "chunks.jsonl")
            .read_text(encoding="utf-8")
            .splitlines()
            if line.strip()
        ]
        embedding_inputs = [
            json.loads(line)
            for line in (documents / "semantic_embedding_input.jsonl")
            .read_text(encoding="utf-8")
            .splitlines()
            if line.strip()
        ]
        tokenized = [
            json.loads(line)
            for line in (documents / "tokenized_chunks.jsonl")
            .read_text(encoding="utf-8")
            .splitlines()
            if line.strip()
        ]
        return TeachingData(
            chunks=chunks,
            embedding_inputs=embedding_inputs,
            tokenized_chunks=tokenized,
            synonyms={},
            graph_nodes={},
            graph_edges=[],
            sqlite_path=workspace / "relational" / "workspace.sqlite3",
            ontology_path=workspace / "graph" / "ontology.ttl",
            instances_path=workspace / "graph" / "instances.ttl",
            shapes_path=workspace / "graph" / "shapes.ttl",
        )

    def table_context(self, workspace_id: str) -> dict[str, Any]:
        workspace = self._path(workspace_id)
        manifest = self.get(workspace_id)
        if manifest.kind != "table" or manifest.status != "ready":
            raise ValueError("该工作区不是可查询的表数据工作区")
        schema = json.loads(
            (workspace / "table" / "schema.json").read_text(
                encoding="utf-8"
            )
        )
        return {
            "manifest": manifest,
            "schema": schema,
            "database_path": workspace / "table" / "database.sqlite3",
        }

    def graph_context(self, workspace_id: str) -> dict[str, Any]:
        workspace = self._path(workspace_id)
        manifest = self.get(workspace_id)
        if manifest.kind != "graph" or manifest.status != "ready":
            raise ValueError("该工作区不是可查询的属性图工作区")
        from .graph_workspace import load_graph_context

        return load_graph_context(workspace, manifest)

    def build_graph(
        self,
        *,
        workspace_id: str,
        name: str,
        source_path: Path,
        update: Callable[[str, int, dict[str, Any] | None], None],
    ) -> WorkspaceInfo:
        from .graph_workspace import build_graph_workspace

        return build_graph_workspace(
            registry_root=self.root,
            workspace_id=workspace_id,
            name=name,
            source_path=source_path,
            update=update,
        )

    def build_documents(
        self,
        *,
        workspace_id: str,
        name: str,
        source_paths: list[Path],
        chunk_size: int,
        chunk_overlap: int,
        update: Callable[[str, int, dict[str, Any] | None], None],
    ) -> WorkspaceInfo:
        building = self.root / f".{workspace_id}.building"
        final = self.root / workspace_id
        building.mkdir(parents=True, exist_ok=False)
        created = _now()
        try:
            update("extract", 12, {"files": len(source_paths)})
            source_dir = building / "source"
            source_dir.mkdir(parents=True)
            extracted: list[dict[str, Any]] = []
            for index, source in enumerate(source_paths, start=1):
                target = source_dir / _safe_upload_name(source.name)
                shutil.copy2(source, target)
                content = _normalize_text(_extract_document(target))
                if not content:
                    raise ValueError(f"未从 {target.name} 提取到文本")
                if len(content) > MAX_DOCUMENT_CHARACTERS_PER_FILE:
                    raise ValueError(
                        f"单文档提取文本不能超过"
                        f"{MAX_DOCUMENT_CHARACTERS_PER_FILE:,}字符：{target.name}"
                    )
                extracted.append(
                    {
                        "doc_id": f"DOC-{index:04d}",
                        "filename": target.name,
                        "text": content,
                        "characters": len(content),
                    }
                )
            total_characters = sum(
                item["characters"] for item in extracted
            )
            if total_characters > MAX_DOCUMENT_CHARACTERS_TOTAL:
                raise ValueError(
                    "本次提取文本总量不能超过"
                    f"{MAX_DOCUMENT_CHARACTERS_TOTAL:,}字符"
                )

            update("chunk", 34, None)
            chunks: list[dict[str, Any]] = []
            for document in extracted:
                parts = chunk_text(
                    document["text"], chunk_size, chunk_overlap
                )
                for index, part in enumerate(parts, start=1):
                    chunks.append(
                        {
                            "chunk_id": (
                                f"{document['doc_id']}-CH-{index:04d}"
                            ),
                            "doc_id": document["doc_id"],
                            "title": document["filename"],
                            "text": part,
                            "source_file": document["filename"],
                            "chunk_index": index,
                            "characters": len(part),
                        }
                    )
            if not chunks:
                raise ValueError("分块结果为空")
            if len(chunks) > MAX_DOCUMENT_CHUNKS:
                raise ValueError(
                    f"自动分块不能超过{MAX_DOCUMENT_CHUNKS:,}块；"
                    "请拆分工作区或增大目标分块长度"
                )

            update("lexical-index", 55, {"chunks": len(chunks)})
            stopwords_path = settings.data_dir / "documents" / "stopwords_zh.txt"
            stopwords = {
                line.strip()
                for line in stopwords_path.read_text(
                    encoding="utf-8-sig"
                ).splitlines()
                if line.strip()
            }
            tokenized: list[dict[str, Any]] = []
            counters: list[Counter[str]] = []
            for chunk in chunks:
                tokens = _tokenize(chunk["text"], stopwords)
                if not tokens:
                    tokens = [chunk["text"][:32]]
                tokenized.append(
                    {"chunk_id": chunk["chunk_id"], "tokens": tokens}
                )
                counters.append(Counter(tokens))
            df: Counter[str] = Counter()
            for counter in counters:
                df.update(counter.keys())
            vocabulary = sorted(df)
            lexical_index = {
                "format_version": 1,
                "document_count": len(chunks),
                "vocabulary_size": len(vocabulary),
                "vocabulary": vocabulary,
                "document_frequency": dict(df),
                "document_lengths": [sum(item.values()) for item in counters],
                "term_counts": [dict(item) for item in counters],
                "tfidf": {
                    "weight": "tf * (log((N+1)/(df+1)) + 1)",
                    "similarity": "cosine",
                },
                "bm25": {"k1": 1.5, "b": 0.75},
            }
            documents_dir = building / "documents"
            _write_jsonl(documents_dir / "chunks.jsonl", chunks)
            _write_jsonl(
                documents_dir / "semantic_embedding_input.jsonl",
                [
                    {"id": item["chunk_id"], "text": item["text"]}
                    for item in chunks
                ],
            )
            _write_jsonl(
                documents_dir / "tokenized_chunks.jsonl", tokenized
            )
            _write_json(documents_dir / "lexical_index.json", lexical_index)
            _write_json(documents_dir / "synonyms.json", {})

            update("vector-index", 72, None)
            placeholder = building / "relational" / "workspace.sqlite3"
            placeholder.parent.mkdir(parents=True, exist_ok=True)
            sqlite3.connect(placeholder).close()
            data = TeachingData(
                chunks=chunks,
                embedding_inputs=[
                    {"id": item["chunk_id"], "text": item["text"]}
                    for item in chunks
                ],
                tokenized_chunks=tokenized,
                synonyms={},
                graph_nodes={},
                graph_edges=[],
                sqlite_path=placeholder,
                ontology_path=building / "graph" / "ontology.ttl",
                instances_path=building / "graph" / "instances.ttl",
                shapes_path=building / "graph" / "shapes.ttl",
            )
            from .document_rag import DocumentRagService

            vector_metadata = DocumentRagService(data).vector_index.ensure(
                force=True
            )
            update(
                "finalize",
                92,
                {
                    "dimensions": vector_metadata["dimensions"],
                    "provider": vector_metadata["provider"],
                },
            )
            manifest = WorkspaceInfo(
                id=workspace_id,
                name=name.strip() or "未命名文档库",
                kind="documents",
                status="ready",
                created_at=created,
                updated_at=_now(),
                source_files=[item["filename"] for item in extracted],
                supported_modes=["tfidf", "bm25", "semantic"],
                statistics={
                    "documents": len(extracted),
                    "chunks": len(chunks),
                    "characters": sum(
                        item["characters"] for item in extracted
                    ),
                    "vocabulary": len(vocabulary),
                    "vectors": vector_metadata["vector_count"],
                    "dimensions": vector_metadata["dimensions"],
                },
                build={
                    "chunk_strategy": "paragraph-sentence-auto",
                    "chunk_size": chunk_size,
                    "chunk_overlap": chunk_overlap,
                    "embedding_provider": vector_metadata["provider"],
                    "artifacts": [
                        "chunks.jsonl",
                        "lexical_index.json",
                        "document_embeddings.faiss",
                        "document_embeddings.meta.json",
                    ],
                },
            )
            _write_json(building / "workspace.json", manifest.model_dump())
            building.replace(final)
            update("completed", 100, manifest.statistics)
            return manifest
        except Exception:
            if building.exists():
                shutil.rmtree(building)
            raise

    def build_table(
        self,
        *,
        workspace_id: str,
        name: str,
        source_path: Path,
        sheet_name: str,
        update: Callable[[str, int, dict[str, Any] | None], None],
    ) -> WorkspaceInfo:
        building = self.root / f".{workspace_id}.building"
        final = self.root / workspace_id
        building.mkdir(parents=True, exist_ok=False)
        created = _now()
        try:
            update("parse-table", 18, None)
            source_dir = building / "source"
            source_dir.mkdir(parents=True)
            target = source_dir / _safe_upload_name(source_path.name)
            shutil.copy2(source_path, target)
            if target.suffix.lower() == ".csv":
                headers, rows, source_detail = _read_csv(target)
            else:
                headers, rows, source_detail = _read_excel(
                    target, sheet_name
                )
            if not rows:
                raise ValueError("表文件没有数据行")

            update("infer-schema", 40, {"rows": len(rows)})
            columns: list[dict[str, Any]] = []
            business_headers = _unique_business_headers(headers)
            for index, (header, original_header) in enumerate(
                business_headers
            ):
                values = [row[index] for row in rows]
                sql_type = _infer_type(values)
                samples: list[Any] = []
                for value in values:
                    if value in (None, "") or value in samples:
                        continue
                    samples.append(value)
                    if len(samples) == 8:
                        break
                columns.append(
                    {
                        "index": index,
                        "source_name": header,
                        "original_source_name": original_header,
                        "sql_name": f"c_{index + 1:03d}",
                        "type": sql_type,
                        "nullable": any(value in (None, "") for value in values),
                        "samples": [str(value)[:100] for value in samples],
                        "distinct_count": len(
                            {str(value) for value in values if value not in (None, "")}
                        ),
                    }
                )

            update("sqlite-import", 62, {"columns": len(columns)})
            table_dir = building / "table"
            table_dir.mkdir(parents=True)
            database_path = table_dir / "database.sqlite3"
            connection = sqlite3.connect(database_path)
            try:
                definitions = ", ".join(
                    f'"{item["sql_name"]}" {item["type"]}'
                    for item in columns
                )
                connection.execute(
                    f'CREATE TABLE "records" ('
                    f'"_row_id" INTEGER PRIMARY KEY, {definitions})'
                )
                column_sql = ", ".join(
                    f'"{item["sql_name"]}"' for item in columns
                )
                placeholders = ", ".join("?" for _ in columns)
                connection.executemany(
                    f'INSERT INTO "records" ({column_sql}) '
                    f"VALUES ({placeholders})",
                    [
                        tuple(
                            _coerce(row[index], columns[index]["type"])
                            for index in range(len(columns))
                        )
                        for row in rows
                    ],
                )
                indexed: list[str] = []
                for item in columns:
                    if (
                        1 < item["distinct_count"] <= min(len(rows) // 2, 1000)
                        and len(indexed) < 5
                    ):
                        index_name = f'idx_records_{item["sql_name"]}'
                        connection.execute(
                            f'CREATE INDEX "{index_name}" ON "records" '
                            f'("{item["sql_name"]}")'
                        )
                        indexed.append(item["sql_name"])
                connection.commit()
            finally:
                connection.close()

            update("profile-table", 84, None)
            schema = {
                "format_version": 1,
                "table_name": "records",
                "source_file": target.name,
                "source_detail": source_detail,
                "row_count": len(rows),
                "columns": columns,
                "indexes": indexed,
                "preview": [
                    {
                        columns[index]["source_name"]: row[index]
                        for index in range(len(columns))
                    }
                    for row in rows[:8]
                ],
            }
            _write_json(table_dir / "schema.json", schema)
            manifest = WorkspaceInfo(
                id=workspace_id,
                name=name.strip() or "未命名表数据库",
                kind="table",
                status="ready",
                created_at=created,
                updated_at=_now(),
                source_files=[target.name],
                supported_modes=["sql"],
                statistics={
                    "rows": len(rows),
                    "columns": len(columns),
                    "indexes": len(indexed),
                    "database_bytes": database_path.stat().st_size,
                },
                build={
                    "table_name": "records",
                    "source_detail": source_detail,
                    "type_inference": True,
                    "artifacts": ["database.sqlite3", "schema.json"],
                },
            )
            _write_json(building / "workspace.json", manifest.model_dump())
            building.replace(final)
            update("completed", 100, manifest.statistics)
            return manifest
        except Exception:
            if building.exists():
                shutil.rmtree(building)
            raise


class IngestionManager:
    STAGES = {
        "documents": [
            ("received", "文件接收"),
            ("extract", "文本提取"),
            ("chunk", "自动分块"),
            ("lexical-index", "TF–IDF / BM25建库"),
            ("vector-index", "Embedding / FAISS建库"),
            ("finalize", "原子发布"),
            ("completed", "完成"),
        ],
        "table": [
            ("received", "文件接收"),
            ("parse-table", "表格解析"),
            ("infer-schema", "字段与类型推断"),
            ("sqlite-import", "SQLite入库与索引"),
            ("profile-table", "数据剖析"),
            ("completed", "完成"),
        ],
        "graph": [
            ("received", "文件接收"),
            ("parse-graph", "JSON / JSONL解析"),
            ("validate-graph", "结构、本体与引用校验"),
            ("sqlite-graph", "SQLite节点边入库"),
            ("node-cards", "节点语义卡构造"),
            ("node-vectors", "Embedding / FAISS锚点索引"),
            ("finalize-graph", "原子发布"),
            ("completed", "完成"),
        ],
    }

    def __init__(self, registry: WorkspaceRegistry) -> None:
        self.registry = registry
        self.staging = settings.runtime_dir / "ingestion"
        self.staging.mkdir(parents=True, exist_ok=True)
        self.executor = ThreadPoolExecutor(
            max_workers=2, thread_name_prefix="rag-ingestion"
        )
        self.jobs: dict[str, dict[str, Any]] = {}
        self.lock = threading.Lock()

    def _new_job(self, kind: str) -> dict[str, Any]:
        job_id = f"job-{uuid.uuid4().hex[:12]}"
        now = _now()
        job = {
            "id": job_id,
            "kind": kind,
            "status": "queued",
            "workspace_id": None,
            "progress": 0,
            "current_stage": "received",
            "stages": [
                {
                    "id": stage_id,
                    "label": label,
                    "status": "pending",
                    "details": {},
                }
                for stage_id, label in self.STAGES[kind]
            ],
            "error": "",
            "created_at": now,
            "updated_at": now,
        }
        job["stages"][0]["status"] = "completed"
        with self.lock:
            self.jobs[job_id] = job
        return job

    def _update(
        self,
        job_id: str,
        stage_id: str,
        progress: int,
        details: dict[str, Any] | None,
    ) -> None:
        with self.lock:
            job = self.jobs[job_id]
            job["status"] = (
                "completed" if stage_id == "completed" else "running"
            )
            job["current_stage"] = stage_id
            job["progress"] = progress
            job["updated_at"] = _now()
            active_index = next(
                index
                for index, item in enumerate(job["stages"])
                if item["id"] == stage_id
            )
            for index, item in enumerate(job["stages"]):
                if index < active_index:
                    item["status"] = "completed"
                elif index == active_index:
                    item["status"] = (
                        "completed" if stage_id == "completed" else "running"
                    )
                    if details:
                        item["details"] = details

    def _fail(self, job_id: str, exc: Exception) -> None:
        with self.lock:
            job = self.jobs[job_id]
            job["status"] = "failed"
            job["error"] = str(exc)
            job["updated_at"] = _now()
            for item in job["stages"]:
                if item["id"] == job["current_stage"]:
                    item["status"] = "failed"

    def _cleanup_staging(self, job_id: str) -> None:
        directory = self.staging / job_id
        if directory.is_dir():
            shutil.rmtree(directory)

    def stage_uploads(
        self, job_id: str, files: list[tuple[str, bytes]], allowed: set[str]
    ) -> list[Path]:
        total = sum(len(content) for _, content in files)
        if total > MAX_UPLOAD_BYTES:
            raise ValueError("单次上传总大小不能超过100MB")
        directory = self.staging / job_id
        directory.mkdir(parents=True, exist_ok=False)
        paths: list[Path] = []
        for index, (name, content) in enumerate(files, start=1):
            if not content:
                raise ValueError(f"文件为空：{name}")
            if len(content) > MAX_FILE_BYTES:
                raise ValueError(f"单个文件不能超过25MB：{name}")
            safe_name = _safe_upload_name(name)
            suffix = Path(safe_name).suffix.lower()
            if suffix not in allowed:
                raise ValueError(f"不支持的文件格式：{suffix or '无扩展名'}")
            path = directory / f"{index:02d}-{safe_name}"
            path.write_bytes(content)
            paths.append(path)
        return paths

    def submit_documents(
        self,
        *,
        name: str,
        files: list[tuple[str, bytes]],
        chunk_size: int,
        chunk_overlap: int,
    ) -> IngestionJobResponse:
        job = self._new_job("documents")
        try:
            paths = self.stage_uploads(job["id"], files, DOCUMENT_EXTENSIONS)
        except Exception:
            self._fail(job["id"], ValueError("文件接收失败"))
            self._cleanup_staging(job["id"])
            raise
        workspace_id = _workspace_id("documents")
        job["workspace_id"] = workspace_id

        def run() -> None:
            try:
                self.registry.build_documents(
                    workspace_id=workspace_id,
                    name=name,
                    source_paths=paths,
                    chunk_size=chunk_size,
                    chunk_overlap=chunk_overlap,
                    update=lambda stage, progress, details: self._update(
                        job["id"], stage, progress, details
                    ),
                )
            except Exception as exc:
                self._fail(job["id"], exc)
            finally:
                self._cleanup_staging(job["id"])

        self.executor.submit(run)
        return self.get(job["id"])

    def submit_table(
        self,
        *,
        name: str,
        file: tuple[str, bytes],
        sheet_name: str,
    ) -> IngestionJobResponse:
        job = self._new_job("table")
        try:
            path = self.stage_uploads(job["id"], [file], TABLE_EXTENSIONS)[0]
        except Exception:
            self._fail(job["id"], ValueError("文件接收失败"))
            self._cleanup_staging(job["id"])
            raise
        workspace_id = _workspace_id("table")
        job["workspace_id"] = workspace_id

        def run() -> None:
            try:
                self.registry.build_table(
                    workspace_id=workspace_id,
                    name=name,
                    source_path=path,
                    sheet_name=sheet_name,
                    update=lambda stage, progress, details: self._update(
                        job["id"], stage, progress, details
                    ),
                )
            except Exception as exc:
                self._fail(job["id"], exc)
            finally:
                self._cleanup_staging(job["id"])

        self.executor.submit(run)
        return self.get(job["id"])

    def submit_graph(
        self,
        *,
        name: str,
        file: tuple[str, bytes],
    ) -> IngestionJobResponse:
        job = self._new_job("graph")
        try:
            path = self.stage_uploads(job["id"], [file], GRAPH_EXTENSIONS)[0]
        except Exception:
            self._fail(job["id"], ValueError("图文件接收失败"))
            self._cleanup_staging(job["id"])
            raise
        workspace_id = _workspace_id("graph")
        job["workspace_id"] = workspace_id

        def run() -> None:
            try:
                self.registry.build_graph(
                    workspace_id=workspace_id,
                    name=name,
                    source_path=path,
                    update=lambda stage, progress, details: self._update(
                        job["id"], stage, progress, details
                    ),
                )
            except Exception as exc:
                self._fail(job["id"], exc)
            finally:
                self._cleanup_staging(job["id"])

        self.executor.submit(run)
        return self.get(job["id"])

    def get(self, job_id: str) -> IngestionJobResponse:
        with self.lock:
            if job_id not in self.jobs:
                raise KeyError("构建任务不存在或后端已重启")
            return IngestionJobResponse.model_validate(
                json.loads(json.dumps(self.jobs[job_id], ensure_ascii=False))
            )


workspace_registry = WorkspaceRegistry()
ingestion_manager = IngestionManager(workspace_registry)
